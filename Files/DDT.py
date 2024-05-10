import openpyxl
path = "D://CREDENCE CLASS//AUTOMATION//automation_concept//openpyxl.xlsx"
workbook = openpyxl.load_workbook(path)
sheet = workbook.active    ## to activate sheet
# workbook.get_sheet_by_name("sheet1")   because we have only one sheet in excel file
rows = sheet.max_row
columns = sheet.max_column

# print(rows)
# print(columns)
for r in range(1, rows+1):
    for c in range(1, columns+1):
        print(sheet.cell(row=r, column=c).value, end="    ")
    print()